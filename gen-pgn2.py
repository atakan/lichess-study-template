#!/usr/bin/env python
import openpyxl
import argparse
import sys


def custom_titlecase(text):
    """
    Converts a string to title case, applying special rules:
    - "I" becomes "ı"
    - "İ" becomes "i"

    Parameters:
        text (str): The input string.

    Returns:
        str: The converted string in title case.
    """
    def custom_capitalize(word):
        if word:
            first_letter = word[0].replace("ı", "I").replace("i", "İ").upper()
            rest = word[1:].replace("I", "ı").replace("İ", "i").lower()
            return first_letter + rest
        return ""

    return " ".join(custom_capitalize(word) for word in text.split())


def capitalize_name(name, lang):
    """
    Properly capitalizes a name based on the given language.

    Parameters:
        name (str): The name to be capitalized.
        lang (str): Language code ('tr' for Turkish, 'en' for English).

    Returns:
        str: Capitalized name.
    """
    if lang == "tr":
        return custom_titlecase(name)
    elif lang == "en":
        return name.title()
    else:
        raise ValueError("Unsupported language code.")


def parse_xlsx(file_path, lang):
    """
    Parses the XLSX file to extract tournament details and game data.

    Parameters:
        file_path (str): Path to the XLSX file.
        lang (str): Language code for capitalization ('tr' or 'en').

    Returns:
        dict: A dictionary with tournament name, player name, player rating, and game details.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        tournament_name = sheet.cell(row=2, column=1).value

        player_name = None
        player_rating = None
        games = []
        table_started = False

        for row in sheet.iter_rows(values_only=True):
            if row[0] == "İsim":
                player_name = capitalize_name(row[4], lang)  # 5th column (zero-indexed 4)
            elif row[0] == "Ulusal rating":
                player_rating = row[4]  # 5th column (zero-indexed 4)
            elif row[0] == "Tur":
                table_started = True  # Start of the important table
                headers = row
                #print(headers)
                round_ind = headers.index('Tur')
                board_ind = headers.index('Masa')
                result_ind = headers.index('Sonuç')
                opp_nam_ind = headers.index('İsim')
                #print("opp_nam_ind", opp_nam_ind)
                opp_rat_ind = opp_int_rat_ind = opp_nat_rat_ind = -1
                if 'Rtg' in headers:
                    ratcase = 1 # rating case 1: we have Rtg
                    opp_rat_ind = headers.index('Rtg')
                elif ('UKD' in headers) and ('ELO' in headers):
                    ratcase = 2 # case 2: we have both UKD and ELO
                    opp_int_rat_ind = headers.index('ELO')
                    opp_nat_rat_ind = headers.index('UKD')                
                elif 'UKD' in headers:
                    ratcase = 3 # case 3: we have UKD but not ELO
                    opp_nat_rat_ind = headers.index('UKD')                
                elif 'ELO' in headers:
                    ratcase = 4 # case 4: we have ELO but not UKD
                    opp_int_rat_ind = headers.index('ELO')
                else :
                    print("something is wrong, rating for the opponent could not be found")
                continue
            elif table_started:
                if all(cell is None for cell in row):  # Empty row, end of table
                    break
                if ratcase == 1:
                    opp_rat = row[opp_rat_ind]
                elif ratcase == 2:
                    if ((row[opp_int_rat_ind] == 0) and
                        (row[opp_nat_rat_ind] == 0)):
                        opp_rat = 0
                    elif row[opp_int_rat_ind] == 0:
                        opp_rat = row[opp_nat_rat_ind]
                    else :
                        opp_rat = row[opp_int_rat_ind]
                elif ratcase == 3:
                    opp_rat = row[opp_nat_rat_ind]
                else :
                    opp_rat = row[opp_int_rat_ind]
                games.append({
                    "round": row[round_ind],
                    "board": row[board_ind],
                    "opponent_name": capitalize_name(row[opp_nam_ind], lang),  
                    "opponent_rating": opp_rat, 
                    "result": row[result_ind].strip(),  # Normalize result by stripping spaces
                })

        return {
            "tournament_name": tournament_name,
            "player_name": player_name,
            "player_rating": player_rating,
            "games": games,
        }
    except Exception as e:
        print(f"Error processing file: {e}")
        sys.exit(1)


def generate_pgn(data, output_file):
    """
    Generates a PGN file from the extracted data.

    Parameters:
        data (dict): The tournament and game data.
        output_file (str): Path to save the PGN file.
    """
    try:
        with open(output_file, "w") as f:
            for game in data["games"]:
                result = game["result"]
                if result == "b 1":
                    pgn_result = "1-0"
                    white = data["player_name"]
                    whiteelo = data["player_rating"]
                    black = game["opponent_name"]
                    blackelo = game["opponent_rating"]
                elif result == "s 1":
                    pgn_result = "0-1"
                    white = game["opponent_name"]
                    whiteelo = game["opponent_rating"]
                    black = data["player_name"]
                    blackelo = data["player_rating"]
                elif result == "b ½":
                    pgn_result = "1/2-1/2"
                    white = data["player_name"]
                    whiteelo = data["player_rating"]
                    black = game["opponent_name"]
                    blackelo = game["opponent_rating"]
                elif result == "s ½":
                    pgn_result = "1/2-1/2"
                    white = game["opponent_name"]
                    whiteelo = game["opponent_rating"]
                    black = data["player_name"]
                    blackelo = data["player_rating"]
                elif result == "b 0":
                    pgn_result = "0-1"
                    white = data["player_name"]
                    whiteelo = data["player_rating"]
                    black = game["opponent_name"]
                    blackelo = game["opponent_rating"]
                elif result == "s 0":
                    pgn_result = "1-0"
                    white = game["opponent_name"]
                    whiteelo = game["opponent_rating"]
                    black = data["player_name"]
                    blackelo = data["player_rating"]
                else:
                    continue  # Skip invalid results

                f.write(f'[Event "{data["tournament_name"]}"]\n')
                f.write(f'[Round "{game["round"]}"]\n')
                f.write(f'[Board "{game["board"]}"]\n')
                f.write(f'[White "{white}"]\n')
                f.write(f'[WhiteUKD "{whiteelo}"]\n')
                f.write(f'[WhiteElo "{whiteelo}"]\n')
                f.write(f'[Black "{black}"]\n')
                f.write(f'[BlackUKD "{blackelo}"]\n')
                f.write(f'[BlackElo "{blackelo}"]\n')
                f.write(f'[Result "{pgn_result}"]\n\n')
                #f.write("1. e4 e5\n\n")
                f.write(f'{pgn_result}\n\n')
        print(f"PGN file generated: {output_file}")
    except Exception as e:
        print(f"Error writing PGN file: {e}")
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description="Parse an XLSX file to generate a PGN file.")
    parser.add_argument("-f", "--file", required=True, help="Path to the XLSX file")
    parser.add_argument("-o", "--output", required=True, help="Path to the output PGN file")
    lang_group = parser.add_mutually_exclusive_group()
    lang_group.add_argument("-t", "--turkish", action="store_true", help="Use Turkish capitalization")
    lang_group.add_argument("-e", "--english", action="store_true", help="Use English capitalization")

    args = parser.parse_args()

    # Default to Turkish if no language is specified
    lang = "tr" if args.turkish or not args.english else "en"

    print("hello")
    data = parse_xlsx(args.file, lang)
    print(data)
    generate_pgn(data, args.output)


if __name__ == "__main__":
    main()
