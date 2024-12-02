import openpyxl
import argparse
import sys


def parse_xlsx(file_path):
    """
    Parses the XLSX file to extract tournament details and game data.

    Parameters:
        file_path (str): Path to the XLSX file.

    Returns:
        dict: A dictionary with tournament name, player name, player rating, and game details.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        tournament_name = sheet.cell(row=2, column=1).value

        player_name = None
        player_rating = None
        games = []
        table_started = False

        for row in sheet.iter_rows(values_only=True):
            if row[0] == "İsim":
                player_name = row[4]  # 5th column (zero-indexed 4)
            elif row[0] == "Ulusal rating":
                player_rating = row[4]  # 5th column (zero-indexed 4)
            elif row[0] == "Tur":
                table_started = True  # Start of the important table
                headers = row
                continue
            elif table_started:
                if all(cell is None for cell in row):  # Empty row, end of table
                    break
                games.append({
                    "round": row[0],
                    "board": row[1],
                    "opponent_name": row[4],  # 5th column (zero-indexed 4)
                    "opponent_rating": row[5],  # 6th column (zero-indexed 5)
                    "result": row[8].strip(),  # Normalize result by stripping spaces
                })

        return {
            "tournament_name": tournament_name,
            "player_name": player_name,
            "player_rating": player_rating,
            "games": games,
        }
    except Exception as e:
        print(f"Error processing file: {e}")
        sys.exit(1)


def generate_pgn(data, output_file):
    """
    Generates a PGN file from the extracted data.

    Parameters:
        data (dict): The tournament and game data.
        output_file (str): Path to save the PGN file.
    """
    try:
        with open(output_file, "w") as f:
            for game in data["games"]:
                result = game["result"]
                if result == "b 1":
                    pgn_result = "1-0"
                    white = data["player_name"]
                    black = game["opponent_name"]
                elif result == "s 1":
                    pgn_result = "0-1"
                    white = game["opponent_name"]
                    black = data["player_name"]
                elif result == "b ½":
                    pgn_result = "1/2-1/2"
                    white = data["player_name"]
                    black = game["opponent_name"]
                elif result == "s ½":
                    pgn_result = "1/2-1/2"
                    white = game["opponent_name"]
                    black = data["player_name"]
                elif result == "b 0":
                    pgn_result = "0-1"
                    white = data["player_name"]
                    black = game["opponent_name"]
                elif result == "s 0":
                    pgn_result = "1-0"
                    white = game["opponent_name"]
                    black = data["player_name"]
                else:
                    continue  # Skip invalid results

                f.write(f'[Event "{data["tournament_name"]}"]\n')
                f.write(f'[Round "{game["round"]}"]\n')
                f.write(f'[Board "{game["board"]}"]\n')
                f.write(f'[White "{white}"]\n')
                f.write(f'[Black "{black}"]\n')
                f.write(f'[Result "{pgn_result}"]\n\n')
                f.write("1. e4 e5\n\n")
        print(f"PGN file generated: {output_file}")
    except Exception as e:
        print(f"Error writing PGN file: {e}")
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description="Parse an XLSX file to generate a PGN file.")
    parser.add_argument("-f", "--file", required=True, help="Path to the XLSX file")
    parser.add_argument("-o", "--output", required=True, help="Path to the output PGN file")
    args = parser.parse_args()

    data = parse_xlsx(args.file)
    generate_pgn(data, args.output)


if __name__ == "__main__":
    main()